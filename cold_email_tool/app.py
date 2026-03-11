"""Cold Email Personalization Tool — Streamlit App.

Minimum input -> Maximum intelligent output.

Usage:
    streamlit run cold_email_tool/app.py
"""

from __future__ import annotations

import csv
import io
import json
import os
import time
import logging
from dataclasses import asdict
from typing import Optional

import streamlit as st

# ---------------------------------------------------------------------------
# Page config — must be first Streamlit call
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Cold Email Tool",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ---------------------------------------------------------------------------
# Imports from the existing backend
# ---------------------------------------------------------------------------
import sys
from pathlib import Path

# Ensure the package root is importable
_pkg_root = str(Path(__file__).resolve().parent.parent)
if _pkg_root not in sys.path:
    sys.path.insert(0, _pkg_root)

from cold_email_tool.core.models import Contact, ResearchBrief, EmailDraft, PipelineResult, AngleCategory
from cold_email_tool.core.config import DEFAULT_SENDER, PipelineConfig, SenderProfile
from cold_email_tool.core.researcher import Researcher
from cold_email_tool.core.email_generator import EmailGenerator
from cold_email_tool.core.enrichment import SparseInput, EnrichmentEngine, EnrichmentResult
from cold_email_tool.core.search import get_search_fn
from cold_email_tool.core.formatter import format_batch_markdown, format_batch_json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Custom CSS for a clean, modern look
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    /* Global font */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }

    /* Hide Streamlit chrome */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}

    /* Tighter spacing */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        max-width: 1100px;
    }

    /* Card style */
    .recipient-card {
        background: #fafafa;
        border: 1px solid #e8e8e8;
        border-radius: 12px;
        padding: 24px;
        margin-bottom: 20px;
    }

    .recipient-card-flagged {
        background: #fffbf0;
        border: 1px solid #f0d68a;
        border-radius: 12px;
        padding: 24px;
        margin-bottom: 20px;
    }

    /* Confidence badges */
    .conf-high {
        background: #e8f5e9;
        color: #2e7d32;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 500;
    }
    .conf-medium {
        background: #fff3e0;
        color: #e65100;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 500;
    }
    .conf-low {
        background: #ffebee;
        color: #c62828;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 500;
    }

    /* Angle pill */
    .angle-pill {
        background: #e3f2fd;
        color: #1565c0;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: 500;
    }

    /* Email body styling */
    .email-body {
        background: white;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        padding: 20px;
        font-size: 0.95rem;
        line-height: 1.6;
        white-space: pre-wrap;
        color: #1a1a1a;
    }

    /* Status badge */
    .status-accepted {
        background: #e8f5e9;
        color: #2e7d32;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
    }
    .status-skipped {
        background: #f5f5f5;
        color: #757575;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
    }
    .status-uncertain {
        background: #fff3e0;
        color: #e65100;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
    }

    /* Step indicator */
    .step-indicator {
        display: flex;
        justify-content: center;
        gap: 32px;
        margin-bottom: 32px;
        padding: 16px 0;
    }
    .step {
        text-align: center;
        color: #bbb;
        font-size: 0.85rem;
        font-weight: 500;
    }
    .step-active {
        text-align: center;
        color: #1a1a1a;
        font-size: 0.85rem;
        font-weight: 600;
    }
    .step-done {
        text-align: center;
        color: #4caf50;
        font-size: 0.85rem;
        font-weight: 500;
    }

    /* Title */
    .app-title {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1a1a1a;
        margin-bottom: 4px;
        letter-spacing: -0.02em;
    }
    .app-subtitle {
        font-size: 0.95rem;
        color: #888;
        margin-bottom: 32px;
    }

    /* Input form */
    .stTextInput > div > div > input {
        border-radius: 8px;
    }
    .stTextArea > div > div > textarea {
        border-radius: 8px;
    }

    /* Button overrides */
    .stButton > button {
        border-radius: 8px;
        font-weight: 500;
        padding: 8px 24px;
    }
    div[data-testid="stForm"] {
        border: 1px solid #e8e8e8;
        border-radius: 12px;
        padding: 24px;
    }

    /* Fact list */
    .fact-item {
        background: #f8f9fa;
        border-left: 3px solid #1565c0;
        padding: 8px 12px;
        margin-bottom: 6px;
        border-radius: 0 6px 6px 0;
        font-size: 0.88rem;
        color: #333;
    }

    /* Subject line chips */
    .subject-chip {
        background: #f0f0f0;
        padding: 4px 12px;
        border-radius: 16px;
        font-size: 0.82rem;
        margin-right: 8px;
        display: inline-block;
        margin-bottom: 4px;
        color: #444;
    }
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Session state initialization
# ---------------------------------------------------------------------------
def init_state():
    defaults = {
        "step": "input",           # input | processing | review | export
        "contacts": [],            # list[SparseInput]
        "enriched": [],            # list[EnrichmentResult]
        "results": [],             # list[PipelineResult]
        "review_statuses": {},     # {idx: "accepted" | "skipped" | "uncertain" | "pending"}
        "edited_emails": {},       # {idx: str}  — user-edited email bodies
        "edited_angles": {},       # {idx: AngleCategory}
        "api_key": os.environ.get("ANTHROPIC_API_KEY", ""),
        "search_provider": "",     # auto-detected
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
ANGLE_LABELS = {
    AngleCategory.AVIATION_ENGINEERING: "Aviation / Engineering",
    AngleCategory.INVESTING_ECON: "Investing / Economic Development",
    AngleCategory.GLOBAL_INFRASTRUCTURE: "Global Infrastructure / China",
    AngleCategory.AI_TOOLS: "AI Tools / Technical Initiative",
    AngleCategory.BUILDER_AMBITION: "Builder Ambition / Intellectual",
}

ANGLE_FROM_LABEL = {v: k for k, v in ANGLE_LABELS.items()}

COLUMN_ALIASES = {
    "full_name": "name", "fullname": "name",
    "linkedin_url": "linkedin", "linkedin_link": "linkedin",
    "twitter_url": "twitter", "twitter_link": "twitter", "x": "twitter",
    "x_url": "twitter", "organization": "company", "org": "company",
    "url": "website", "site": "website", "web": "website",
    "note": "notes", "comment": "notes", "comments": "notes",
}

CANONICAL_COLS = {"name", "email", "linkedin", "company", "website", "twitter", "notes"}


def normalize_columns(df_columns: list[str]) -> dict[str, str]:
    """Map uploaded column names to canonical names."""
    mapping = {}
    for col in df_columns:
        lower = col.strip().lower().replace(" ", "_")
        if lower in CANONICAL_COLS:
            mapping[col] = lower
        elif lower in COLUMN_ALIASES:
            mapping[col] = COLUMN_ALIASES[lower]
    return mapping


def confidence_badge(score: float) -> str:
    if score >= 0.8:
        return f'<span class="conf-high">High ({score:.0%})</span>'
    elif score >= 0.5:
        return f'<span class="conf-medium">Medium ({score:.0%})</span>'
    else:
        return f'<span class="conf-low">Low ({score:.0%})</span>'


def status_badge(status: str) -> str:
    css = {"accepted": "status-accepted", "skipped": "status-skipped", "uncertain": "status-uncertain"}
    return f'<span class="{css.get(status, "")}">{status}</span>'


def step_indicator(current: str):
    steps = [("input", "Input"), ("processing", "Research"), ("review", "Review"), ("export", "Export")]
    idx = next(i for i, (k, _) in enumerate(steps) if k == current)
    parts = []
    for i, (key, label) in enumerate(steps):
        num = i + 1
        if i < idx:
            parts.append(f'<div class="step-done">{num}. {label}</div>')
        elif i == idx:
            parts.append(f'<div class="step-active">{num}. {label}</div>')
        else:
            parts.append(f'<div class="step">{num}. {label}</div>')
    st.markdown(f'<div class="step-indicator">{"".join(parts)}</div>', unsafe_allow_html=True)


def get_pipeline(api_key: str):
    """Initialize the pipeline with real search if available."""
    search_fn = get_search_fn()
    config = PipelineConfig(anthropic_api_key=api_key)
    return config, search_fn


# ---------------------------------------------------------------------------
# Step 1: INPUT
# ---------------------------------------------------------------------------
def render_input():
    st.markdown('<div class="app-title">Cold Email Tool</div>', unsafe_allow_html=True)
    st.markdown('<div class="app-subtitle">Minimum input. Maximum personalization.</div>', unsafe_allow_html=True)

    step_indicator("input")

    # API key check
    api_key = st.session_state.api_key
    if not api_key:
        api_key = st.text_input(
            "Anthropic API Key",
            type="password",
            placeholder="sk-ant-...",
            help="Required. Set ANTHROPIC_API_KEY env var to skip this.",
        )
        if api_key:
            st.session_state.api_key = api_key

    # Search provider status
    search_fn = get_search_fn()
    has_search = search_fn.__name__ != "_stub_search"
    if has_search:
        st.success(f"Search: {search_fn.__name__} connected", icon=None)
    else:
        st.info(
            "No search API configured. Set `BRAVE_API_KEY`, `TAVILY_API_KEY`, or `SERPAPI_KEY` "
            "for live research. The tool will still work using Claude's knowledge.",
            icon=None,
        )

    tab_single, tab_batch, tab_upload = st.tabs(["Single Person", "Batch Entry", "Upload File"])

    # --- Single entry ---
    with tab_single:
        st.markdown("Enter a name and any one identifier. That's enough.")
        with st.form("single_form", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("Name *", placeholder="e.g. Patrick Collison")
            with col2:
                identifier = st.text_input(
                    "Email, LinkedIn, company, website, or X",
                    placeholder="e.g. Stripe, or a URL",
                )
            notes = st.text_input("Notes (optional)", placeholder="e.g. Interested in progress studies")
            submitted = st.form_submit_button("Add Person", use_container_width=True)

        if submitted and name.strip():
            sparse = _parse_single_input(name.strip(), identifier.strip(), notes.strip())
            st.session_state.contacts.append(sparse)
            st.rerun()

    # --- Batch entry ---
    with tab_batch:
        st.markdown("Paste multiple people, one per line. Format: `Name, identifier, notes` (comma-separated).")
        batch_text = st.text_area(
            "Batch input",
            height=200,
            placeholder="Patrick Collison, Stripe\nGwynne Shotwell, SpaceX\nNat Friedman, https://nat.org, AI investor",
            label_visibility="collapsed",
        )
        if st.button("Add All", key="batch_add", use_container_width=True):
            if batch_text.strip():
                for line in batch_text.strip().split("\n"):
                    parts = [p.strip() for p in line.split(",")]
                    if parts and parts[0]:
                        name = parts[0]
                        ident = parts[1] if len(parts) > 1 else ""
                        note = parts[2] if len(parts) > 2 else ""
                        sparse = _parse_single_input(name, ident, note)
                        st.session_state.contacts.append(sparse)
                st.rerun()

    # --- File upload ---
    with tab_upload:
        st.markdown("Upload a CSV or Excel file. Only `name` is required — everything else is optional.")
        st.markdown(
            '**Supported columns:** `name`, `email`, `linkedin`, `company`, `website`, `twitter`, `notes`'
        )
        uploaded = st.file_uploader(
            "Choose file",
            type=["csv", "xlsx", "xls"],
            label_visibility="collapsed",
        )
        if uploaded:
            _handle_file_upload(uploaded)

    # --- Show queued contacts ---
    if st.session_state.contacts:
        st.markdown("---")
        st.markdown(f"### Queued ({len(st.session_state.contacts)} people)")

        for i, s in enumerate(st.session_state.contacts):
            cols = st.columns([4, 2, 2, 1])
            with cols[0]:
                st.markdown(f"**{s.name}**")
            with cols[1]:
                detail = s.company or s.email or s.linkedin or s.website or s.twitter or ""
                if detail:
                    st.caption(detail[:40])
            with cols[2]:
                if s.notes:
                    st.caption(s.notes[:30])
            with cols[3]:
                if st.button("x", key=f"rm_{i}", help="Remove"):
                    st.session_state.contacts.pop(i)
                    st.rerun()

        st.markdown("")
        col_go, col_clear = st.columns([3, 1])
        with col_go:
            if st.session_state.api_key:
                if st.button("Research & Generate Emails", type="primary", use_container_width=True):
                    st.session_state.step = "processing"
                    st.rerun()
            else:
                st.warning("Enter your Anthropic API key above to proceed.")
        with col_clear:
            if st.button("Clear All", use_container_width=True):
                st.session_state.contacts = []
                st.rerun()


def _parse_single_input(name: str, identifier: str, notes: str = "") -> SparseInput:
    """Parse a free-form identifier into the right SparseInput field."""
    sparse = SparseInput(name=name, notes=notes)

    if not identifier:
        return sparse

    ident = identifier.strip()

    # Email
    if "@" in ident and "." in ident.split("@")[-1]:
        sparse.email = ident
    # LinkedIn
    elif "linkedin.com" in ident.lower():
        sparse.linkedin = ident if ident.startswith("http") else f"https://{ident}"
    # Twitter/X
    elif "twitter.com" in ident.lower() or "x.com" in ident.lower():
        sparse.twitter = ident if ident.startswith("http") else f"https://{ident}"
    elif ident.startswith("@"):
        sparse.twitter = f"https://x.com/{ident.lstrip('@')}"
    # URL
    elif ident.startswith("http://") or ident.startswith("https://"):
        sparse.website = ident
    elif "." in ident and " " not in ident and len(ident) > 4:
        # Looks like a domain
        sparse.website = f"https://{ident}"
    else:
        # Assume it's a company name
        sparse.company = ident

    return sparse


def _handle_file_upload(uploaded):
    """Parse an uploaded CSV or Excel file into SparseInputs."""
    try:
        if uploaded.name.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                st.error("Empty file.")
                return
            headers = [str(h).strip() if h else "" for h in rows[0]]
            data_rows = rows[1:]
        else:
            content = uploaded.read().decode("utf-8")
            reader = csv.reader(io.StringIO(content))
            all_rows = list(reader)
            if not all_rows:
                st.error("Empty file.")
                return
            headers = [h.strip() for h in all_rows[0]]
            data_rows = all_rows[1:]

        col_map = normalize_columns(headers)

        if not col_map:
            st.error(f"No recognized columns found. Got: {headers}")
            return

        count = 0
        for row in data_rows:
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j] in col_map:
                    row_dict[col_map[headers[j]]] = str(val).strip() if val else ""

            name = row_dict.get("name", "")
            if not name:
                continue

            sparse = SparseInput(
                name=name,
                email=row_dict.get("email", ""),
                linkedin=row_dict.get("linkedin", ""),
                company=row_dict.get("company", ""),
                website=row_dict.get("website", ""),
                twitter=row_dict.get("twitter", ""),
                notes=row_dict.get("notes", ""),
            )
            st.session_state.contacts.append(sparse)
            count += 1

        st.success(f"Loaded {count} contacts from {uploaded.name}")
        st.rerun()

    except Exception as e:
        st.error(f"Failed to parse file: {e}")


# ---------------------------------------------------------------------------
# Step 2: PROCESSING
# ---------------------------------------------------------------------------
def render_processing():
    st.markdown('<div class="app-title">Researching & Writing</div>', unsafe_allow_html=True)
    st.markdown('<div class="app-subtitle">Enriching profiles, researching recipients, generating emails.</div>', unsafe_allow_html=True)
    step_indicator("processing")

    contacts = st.session_state.contacts
    api_key = st.session_state.api_key
    search_fn = get_search_fn()

    config = PipelineConfig(anthropic_api_key=api_key)
    enrichment = EnrichmentEngine(api_key, model=config.model, search_fn=search_fn)
    researcher = Researcher(config, DEFAULT_SENDER, search_fn=search_fn)
    generator = EmailGenerator(config, DEFAULT_SENDER)

    progress = st.progress(0.0)
    status = st.empty()
    results = []

    total = len(contacts)
    for i, sparse in enumerate(contacts):
        pct = i / total

        # Phase 1: Enrich
        status.markdown(f"**{i+1}/{total}** — Enriching **{sparse.name}**...")
        progress.progress(pct + 0.1 / total)

        try:
            enriched = enrichment.enrich(sparse)
        except Exception as e:
            logger.error("Enrichment failed for %s: %s", sparse.name, e)
            enriched = EnrichmentResult(
                original_input=sparse,
                contact=Contact(
                    full_name=sparse.name,
                    company=sparse.company,
                    website=sparse.website,
                    linkedin=sparse.linkedin,
                    twitter=sparse.twitter,
                ),
                identity_confidence=0.1,
                ambiguity_note=f"Enrichment error: {e}",
            )

        # Phase 2: Research
        status.markdown(f"**{i+1}/{total}** — Researching **{enriched.contact.full_name}**...")
        progress.progress(pct + 0.4 / total)

        try:
            brief = researcher.research_contact(enriched.contact)
        except Exception as e:
            logger.error("Research failed for %s: %s", enriched.contact.full_name, e)
            brief = ResearchBrief(contact=enriched.contact, confidence_score=0.1)

        # Phase 3: Generate email
        status.markdown(f"**{i+1}/{total}** — Drafting email for **{enriched.contact.full_name}**...")
        progress.progress(pct + 0.7 / total)

        try:
            draft = generator.generate(brief)
        except Exception as e:
            logger.error("Generation failed for %s: %s", enriched.contact.full_name, e)
            draft = EmailDraft(
                recipient=enriched.contact,
                research_brief=brief,
                body="[Generation failed — manual draft required]",
                angle_used=brief.background_emphasis,
            )

        # Build result
        flagged = (
            brief.confidence_score < config.low_confidence_threshold
            or enriched.identity_confidence < 0.5
            or not brief.is_sufficient
        )
        review_reason = ""
        if enriched.identity_confidence < 0.5:
            review_reason = f"Low identity confidence ({enriched.identity_confidence:.0%}). "
        if brief.confidence_score < config.low_confidence_threshold:
            review_reason += f"Low research confidence ({brief.confidence_score:.0%}). "
        if enriched.ambiguity_note:
            review_reason += enriched.ambiguity_note

        result = PipelineResult(
            contact=enriched.contact,
            research_brief=brief,
            email_draft=draft,
            flagged_for_review=flagged,
            review_reason=review_reason.strip(),
        )
        results.append(result)

        progress.progress((i + 1) / total)

    st.session_state.results = results
    st.session_state.review_statuses = {i: "pending" for i in range(len(results))}
    st.session_state.step = "review"
    status.markdown("Done. Redirecting to review...")
    time.sleep(0.5)
    st.rerun()


# ---------------------------------------------------------------------------
# Step 3: REVIEW
# ---------------------------------------------------------------------------
def render_review():
    st.markdown('<div class="app-title">Review Drafts</div>', unsafe_allow_html=True)
    st.markdown('<div class="app-subtitle">Review, edit, and approve each email before exporting.</div>', unsafe_allow_html=True)
    step_indicator("review")

    results = st.session_state.results
    if not results:
        st.warning("No results to review. Go back to input.")
        if st.button("Back to Input"):
            st.session_state.step = "input"
            st.rerun()
        return

    # Summary bar
    total = len(results)
    flagged = sum(1 for r in results if r.flagged_for_review)
    accepted = sum(1 for s in st.session_state.review_statuses.values() if s == "accepted")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total", total)
    col2.metric("Flagged", flagged)
    col3.metric("Accepted", accepted)
    col4.metric("Remaining", total - accepted - sum(1 for s in st.session_state.review_statuses.values() if s == "skipped"))

    st.markdown("---")

    # Render each result
    for idx, result in enumerate(results):
        _render_result_card(idx, result)

    # Navigation
    st.markdown("---")
    col_back, col_export = st.columns([1, 3])
    with col_back:
        if st.button("Back to Input", use_container_width=True):
            st.session_state.step = "input"
            st.rerun()
    with col_export:
        if st.button("Export Results", type="primary", use_container_width=True):
            st.session_state.step = "export"
            st.rerun()


def _render_result_card(idx: int, result: PipelineResult):
    """Render a single recipient card in the review UI."""
    c = result.contact
    b = result.research_brief
    e = result.email_draft
    status = st.session_state.review_statuses.get(idx, "pending")

    card_class = "recipient-card-flagged" if result.flagged_for_review else "recipient-card"

    # Header
    name_display = c.full_name
    role_display = ""
    if c.role and c.company:
        role_display = f"{c.role} at {c.company}"
    elif c.company:
        role_display = c.company
    elif c.role:
        role_display = c.role

    conf = confidence_badge(b.confidence_score)
    angle_label = ANGLE_LABELS.get(
        st.session_state.edited_angles.get(idx, e.angle_used),
        e.angle_used.value.replace("_", " ").title()
    )
    angle_html = f'<span class="angle-pill">{angle_label}</span>'

    status_html = ""
    if status != "pending":
        status_html = f" {status_badge(status)}"

    st.markdown(
        f'<div class="{card_class}">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">'
        f'<div><strong style="font-size:1.1rem;">{name_display}</strong>'
        f'{f" <span style=&quot;color:#666;font-size:0.9rem;&quot;>— {role_display}</span>" if role_display else ""}'
        f'{status_html}</div>'
        f'<div>{conf} {angle_html}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # Warning
    if result.flagged_for_review and result.review_reason:
        st.markdown(
            f'<div style="background:#fff8e1;border-left:3px solid #ffa000;padding:8px 12px;border-radius:0 6px 6px 0;font-size:0.85rem;margin-bottom:12px;color:#6d4c00;">'
            f'{result.review_reason}</div>',
            unsafe_allow_html=True,
        )

    # Key facts
    if b.specific_facts:
        for fact in b.specific_facts:
            if fact and "INSUFFICIENT" not in fact.upper():
                st.markdown(f'<div class="fact-item">{fact}</div>', unsafe_allow_html=True)

    # Email body
    st.markdown('<div style="margin-top:12px;"></div>', unsafe_allow_html=True)
    current_body = st.session_state.edited_emails.get(idx, e.body)
    st.markdown(f'<div class="email-body">{current_body}</div>', unsafe_allow_html=True)

    # Subject lines
    if e.subject_lines:
        subjs = " ".join(f'<span class="subject-chip">{s}</span>' for s in e.subject_lines)
        st.markdown(f'<div style="margin-top:8px;">{subjs}</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # Controls
    cols = st.columns([1, 1, 1, 1, 2])
    with cols[0]:
        if st.button("Accept", key=f"accept_{idx}", use_container_width=True):
            st.session_state.review_statuses[idx] = "accepted"
            st.rerun()
    with cols[1]:
        if st.button("Skip", key=f"skip_{idx}", use_container_width=True):
            st.session_state.review_statuses[idx] = "skipped"
            st.rerun()
    with cols[2]:
        if st.button("Uncertain", key=f"uncertain_{idx}", use_container_width=True):
            st.session_state.review_statuses[idx] = "uncertain"
            st.rerun()
    with cols[3]:
        angle_options = list(ANGLE_LABELS.values())
        current_angle = ANGLE_LABELS.get(
            st.session_state.edited_angles.get(idx, e.angle_used),
            angle_options[-1]
        )
        new_angle = st.selectbox(
            "Angle",
            angle_options,
            index=angle_options.index(current_angle) if current_angle in angle_options else 0,
            key=f"angle_{idx}",
            label_visibility="collapsed",
        )
        if new_angle != current_angle:
            st.session_state.edited_angles[idx] = ANGLE_FROM_LABEL[new_angle]
    with cols[4]:
        with st.expander("Edit email"):
            edited = st.text_area(
                "Edit body",
                value=current_body,
                height=150,
                key=f"edit_{idx}",
                label_visibility="collapsed",
            )
            if edited != current_body:
                st.session_state.edited_emails[idx] = edited

    st.markdown("")


# ---------------------------------------------------------------------------
# Step 4: EXPORT
# ---------------------------------------------------------------------------
def render_export():
    st.markdown('<div class="app-title">Export Results</div>', unsafe_allow_html=True)
    st.markdown('<div class="app-subtitle">Download your personalized emails in any format.</div>', unsafe_allow_html=True)
    step_indicator("export")

    results = st.session_state.results
    if not results:
        st.warning("No results to export.")
        if st.button("Back to Input"):
            st.session_state.step = "input"
            st.rerun()
        return

    # Apply edits to results
    export_data = _build_export_data(results)

    # Summary
    total = len(export_data)
    accepted = sum(1 for r in export_data if r["review_status"] == "accepted")
    skipped = sum(1 for r in export_data if r["review_status"] == "skipped")

    col1, col2, col3 = st.columns(3)
    col1.metric("Total", total)
    col2.metric("Accepted", accepted)
    col3.metric("Skipped", skipped)

    st.markdown("---")

    # Filter option
    export_filter = st.radio(
        "What to export",
        ["All", "Accepted only", "Accepted + Uncertain"],
        horizontal=True,
    )

    if export_filter == "Accepted only":
        export_data = [r for r in export_data if r["review_status"] == "accepted"]
    elif export_filter == "Accepted + Uncertain":
        export_data = [r for r in export_data if r["review_status"] in ("accepted", "uncertain")]

    if not export_data:
        st.info("No results match the filter.")
        return

    st.markdown(f"**{len(export_data)} emails** ready for export.")

    # Format options
    col_csv, col_xlsx, col_json, col_md = st.columns(4)

    with col_csv:
        csv_str = _to_csv(export_data)
        st.download_button(
            "Download CSV",
            data=csv_str,
            file_name="cold_emails.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_xlsx:
        xlsx_bytes = _to_xlsx(export_data)
        if xlsx_bytes:
            st.download_button(
                "Download XLSX",
                data=xlsx_bytes,
                file_name="cold_emails.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("Install `openpyxl` for XLSX export.")

    with col_json:
        json_str = json.dumps(export_data, indent=2)
        st.download_button(
            "Download JSON",
            data=json_str,
            file_name="cold_emails.json",
            mime="application/json",
            use_container_width=True,
        )

    with col_md:
        md_str = _to_markdown(export_data)
        st.download_button(
            "Download Markdown",
            data=md_str,
            file_name="cold_emails.md",
            mime="text/markdown",
            use_container_width=True,
        )

    # Preview
    st.markdown("---")
    st.markdown("### Preview")
    for row in export_data:
        with st.expander(f"{row['input_name']} → {row['resolved_name']}"):
            st.markdown(f"**{row['role']}** at **{row['company']}**")
            st.markdown(f"Confidence: {row['confidence']} | Angle: {row['chosen_angle']}")
            if row.get("warning_flag"):
                st.warning(row["warning_flag"])
            st.markdown("**Subject lines:**")
            for s in [row.get("subject_line_1"), row.get("subject_line_2"), row.get("subject_line_3")]:
                if s:
                    st.markdown(f"- {s}")
            st.markdown("**Email:**")
            st.text(row["email_body"])

    # Back
    st.markdown("---")
    col_back, col_new = st.columns(2)
    with col_back:
        if st.button("Back to Review", use_container_width=True):
            st.session_state.step = "review"
            st.rerun()
    with col_new:
        if st.button("Start New Batch", use_container_width=True):
            for key in ["contacts", "enriched", "results", "review_statuses", "edited_emails", "edited_angles"]:
                st.session_state[key] = [] if isinstance(st.session_state[key], list) else {}
            st.session_state.step = "input"
            st.rerun()


def _build_export_data(results: list[PipelineResult]) -> list[dict]:
    """Build the flat export data with all requested columns."""
    rows = []
    for idx, r in enumerate(results):
        # Get potentially edited values
        body = st.session_state.edited_emails.get(idx, r.email_draft.body)
        angle = st.session_state.edited_angles.get(idx, r.email_draft.angle_used)
        status = st.session_state.review_statuses.get(idx, "pending")

        subjs = r.email_draft.subject_lines + ["", "", ""]  # pad

        sparse = st.session_state.contacts[idx] if idx < len(st.session_state.contacts) else None
        input_name = sparse.name if sparse else r.contact.full_name

        rows.append({
            "input_name": input_name,
            "resolved_name": r.contact.full_name,
            "email": sparse.email if sparse else "",
            "linkedin": r.contact.linkedin,
            "company": r.contact.company,
            "role": r.contact.role,
            "chosen_angle": ANGLE_LABELS.get(angle, angle.value if hasattr(angle, 'value') else str(angle)),
            "confidence": f"{r.research_brief.confidence_score:.0%}",
            "subject_line_1": subjs[0],
            "subject_line_2": subjs[1],
            "subject_line_3": subjs[2],
            "email_body": body,
            "review_status": status,
            "warning_flag": r.review_reason if r.flagged_for_review else "",
        })
    return rows


def _to_csv(data: list[dict]) -> str:
    if not data:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()


def _to_xlsx(data: list[dict]) -> Optional[bytes]:
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cold Emails"

    if not data:
        return None

    # Headers
    headers = list(data[0].keys())
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for i, row in enumerate(data, 2):
        for j, key in enumerate(headers, 1):
            ws.cell(row=i, column=j, value=row.get(key, ""))

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_markdown(data: list[dict]) -> str:
    lines = ["# Cold Email Drafts\n"]
    for row in data:
        lines.append(f"## {row['resolved_name']}")
        if row["role"] or row["company"]:
            lines.append(f"**{row['role']}** at **{row['company']}**\n")
        lines.append(f"- Angle: {row['chosen_angle']}")
        lines.append(f"- Confidence: {row['confidence']}")
        lines.append(f"- Status: {row['review_status']}")
        if row.get("warning_flag"):
            lines.append(f"- Warning: {row['warning_flag']}")
        lines.append("")
        for i in range(1, 4):
            s = row.get(f"subject_line_{i}", "")
            if s:
                lines.append(f"**Subject {i}:** {s}")
        lines.append(f"\n{row['email_body']}\n")
        lines.append("---\n")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main routing
# ---------------------------------------------------------------------------
def main():
    step = st.session_state.step
    if step == "input":
        render_input()
    elif step == "processing":
        render_processing()
    elif step == "review":
        render_review()
    elif step == "export":
        render_export()
    else:
        st.session_state.step = "input"
        st.rerun()


if __name__ == "__main__":
    main()
else:
    main()
